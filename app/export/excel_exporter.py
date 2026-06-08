"""Excel导出模块 - 将发票数据导出为按日期命名的Excel文件"""

import os
from datetime import date
from typing import Optional
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class ExcelExporter:
    """发票数据导出为Excel"""

    # 表头定义
    HEADERS = [
        '文件名', '发票号码', '开票日期',
        '购方名称', '购方纳税人识别号',
        '销方名称', '销方纳税人识别号',
        '开票项目', '开票金额',
    ]

    # 列宽设置
    COLUMN_WIDTHS = {
        '文件名': 40,
        '发票号码': 20,
        '开票日期': 15,
        '购方名称': 30,
        '购方纳税人识别号': 25,
        '销方名称': 30,
        '销方纳税人识别号': 25,
        '开票项目': 50,
        '开票金额': 15,
    }

    @staticmethod
    def export(data_list: list[dict], output_dir: str,
               custom_name: Optional[str] = None) -> str:
        """导出发票数据到Excel文件

        Args:
            data_list: 发票数据字典列表
            output_dir: 输出目录
            custom_name: 自定义文件名（不含扩展名），默认按日期自动生成

        Returns:
            导出的Excel文件路径
        """
        # 生成文件名
        if custom_name:
            file_name = f"{custom_name}.xlsx"
        else:
            today = date.today()
            file_name = f"发票数据_{today.strftime('%Y-%m-%d')}.xlsx"

        file_path = os.path.join(output_dir, file_name)

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "发票数据"

        # 写入表头
        header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4',
                                  fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center',
                                     wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for col_idx, header in enumerate(ExcelExporter.HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

            # 设置列宽
            width = ExcelExporter.COLUMN_WIDTHS.get(header, 20)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # 写入数据行
        data_font = Font(name='微软雅黑', size=10)
        data_alignment = Alignment(vertical='center', wrap_text=True)

        for row_idx, data in enumerate(data_list, 2):
            for col_idx, header in enumerate(ExcelExporter.HEADERS, 1):
                value = data.get(header, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

        # 冻结首行
        ws.freeze_panes = 'A2'

        # 自动筛选
        ws.auto_filter.ref = ws.dimensions

        # 保存
        wb.save(file_path)
        return file_path
